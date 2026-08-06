import re
import sys
import os
import json
import csv
import datetime
import openpyxl

# Reconfigure output to support UTF-8 characters on Windows console
sys.stdout.reconfigure(encoding='utf-8')

# Paths
EXCEL_PATH = "LabFlow – Dữ liệu Report cuối buổi.xlsx"
JSON_PATH = "src/data/initialData.json"
PITCHING_CSV_PATH = "../Danh_Sach_Thanh_Vien_Top_Rank_K3_K4.csv"

def clean_student_name(name):
    """
    Cleans student names by removing trailing or leading punctuation,
    extra spaces, and common activity suffixes accidentally appended to the name.
    """
    if not name:
        return ""
    name = name.strip()
    # Remove leading hyphens, bullets or symbols
    name = re.sub(r'^[-\s+*•]+', '', name)
    # Remove trailing plus signs or spaces
    name = re.sub(r'[+\s]+$', '', name)
    # Strip common activity keywords that were accidentally treated as part of the student's name
    activity_keywords = [
        "phát biểu", "giơ tay phát biểu", "phát biểu xây dựng bài", 
        "phát biểu trong giờ", "lên demo lab", "demo bài lab", "demo lab", 
        "demo ý tưởng", "top 10 kahoot", "trong top 10 kahoots", "bonus quiz", 
        "trả lời câu hỏi lý thuyết", "tham gia hỏi đáp qa", 
        "hoàn thành bài sớm và hỗ trợ các bạn", "hỗ trợ các nhóm khác hoàn thành bài lab",
        "hỏi đáp qa", "demo sản phẩm bàn", "demo"
    ]
    # Regex to match any of the keywords as word boundary and strip everything after
    keyword_pattern = r'(?i)\b(' + '|'.join(activity_keywords) + r')\b.*$'
    name = re.sub(keyword_pattern, '', name)
    
    # Final clean of trailing punctuation and spaces
    name = name.strip()
    name = re.sub(r'[-\s+*•]+$', '', name)
    
    # Strip leading special chars: double quotes, single quotes, parens, brackets
    name = re.sub(r'^["\x27(\[{]+', '', name).strip()
    
    # Correct capitalization for names (Vietnamese safe)
    words = name.split()
    result = []
    for w in words:
        if w:
            result.append(w[0].upper() + w[1:] if len(w) > 1 else w.upper())
    
    return ' '.join(result)

def extract_student_id_and_raw(line):
    """
    Extracts the unique student ID by looking for 2A... codes or 3-5 digit suffixes.
    Standardizes them to the 2A20260 + 4-digit code format.
    """
    # Matches patterns like 2A202601079, 2A20601302, T158, T013, or standalone 3-5 digits
    match_id = re.search(r'([2T][A0-9]*\d{3,5}|\b\d{3,5}\b)', line)
    if not match_id:
        return None, None

    raw_id = match_id.group(1)
    digits = re.sub(r'\D', '', raw_id)
    if len(digits) < 3:
        return None, None
    suffix = digits[-4:].zfill(4)
    student_id = f"2A20260{suffix}"
    return student_id, raw_id

def parse_raw_line(line):
    """
    Parses a raw log line to extract Student ID, Name, Points, and Reason.
    """
    line = line.strip()
    if not line:
        return None

    student_id, raw_id = extract_student_id_and_raw(line)
    if not student_id:
        return None # Not a valid student record

    # Extract points (looks for standalone single digits 1, 2, 3, 5 or with a '+' prefix)
    points = 1
    # Remove raw ID from line first to avoid matching parts of the ID as points
    line_without_id = line.replace(raw_id, ' [ID] ')
    points_match = re.search(r'(?:\b|\s)\+?([1235])(?:\b|\s)', line_without_id)
    if points_match:
        points = int(points_match.group(1))

    # Categorize and standardize reason
    reason = "Phát biểu"
    if "kahoot" in line.lower() or "quiz" in line.lower():
        reason = "Top Kahoot / Quiz"
    elif "demo" in line.lower():
        reason = "Demo Lab / Ý tưởng"
    elif "hỗ trợ" in line.lower() or "hoàn thành bài sớm" in line.lower():
        reason = "Hỗ trợ học viên khác"
    elif "phản biện" in line.lower():
        reason = "Phát biểu phản biện"
    elif "lý thuyết" in line.lower() or "hỏi đáp" in line.lower() or "qa" in line.lower():
        reason = "Phát biểu lý thuyết / Q&A"
    
    # Extract student name by cleaning the remaining line contents
    cleaned_line = line.replace(raw_id, '')
    if points_match:
        cleaned_line = cleaned_line.replace(points_match.group(0), '')
        
    parts = [p.strip() for p in re.split(r'[\t\-–—:,]', cleaned_line) if p.strip()]
    
    name_candidate = ""
    # Heuristic 1: Find a clean-looking part that is long enough and contains no keywords
    for p in parts:
        cleaned_p = clean_student_name(p)
        if len(cleaned_p) > 4 and not any(k in p.lower() for k in ["phát biểu", "demo", "kahoot", "quiz", "hỗ trợ", "bonus"]):
            name_candidate = cleaned_p
            break
            
    # Heuristic 2: Fall back to first non-empty cleaned part
    if not name_candidate and parts:
        for p in parts:
            cleaned_p = clean_student_name(p)
            if len(cleaned_p) > 2:
                name_candidate = cleaned_p
                break
                
    # Heuristic 3: Clean the entire line
    if not name_candidate:
        name_candidate = clean_student_name(cleaned_line)

    return {
        "student_id": student_id,
        "student_name": name_candidate,
        "points": points,
        "reason": reason,
        "raw_line": line
    }

def main():
    print("=== Starting LabScore Data Cleanup ===")

    # 1. Load Student Metadata Cache from current initialData.json if it exists
    student_cache = {}
    if os.path.exists(JSON_PATH):
        try:
            with open(JSON_PATH, "r", encoding="utf-8") as f:
                old_data = json.load(f)
                for s in old_data.get("students", []):
                    # Cache historical student details (role, project name, etc.)
                    # Only cache if it is a clean, official student ID
                    if s["id"].startswith("2A20260"):
                        student_cache[s["id"]] = {
                            "name": clean_student_name(s["name"]),
                            "course": s.get("course", "Khóa 3"),
                            "team_code": s.get("team_code", "N/A"),
                            "project_name": s.get("project_name", "Chưa có dự án"),
                            "role": s.get("role", "Thành viên"),
                            "ai_score": s.get("ai_score", 0),
                            "class_name": s.get("class_name", "C401"),
                            "trend": s.get("trend", "flat"),
                            "streak": s.get("streak", 0)
                        }
            print(f"Loaded {len(student_cache)} official students from initialData.json cache.")
        except Exception as e:
            print(f"Warning: Could not read existing initialData.json: {e}")

    # 2. Enrich/Load Pitching Day top rank students from CSV
    pitching_count = 0
    if os.path.exists(PITCHING_CSV_PATH):
        try:
            with open(PITCHING_CSV_PATH, mode='r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    student_id = row.get("Mã học viên", "").strip()
                    if student_id:
                        student_cache[student_id] = {
                            "name": clean_student_name(row.get("Họ và tên", "")),
                            "course": row.get("Khóa", "Khóa 3").strip(),
                            "team_code": row.get("Mã nhóm", "N/A").strip(),
                            "project_name": row.get("Tên đề tài / Dự án", "Chưa có dự án").strip(),
                            "role": row.get("Vai trò trong nhóm", "Thành viên").strip(),
                            "ai_score": int(row.get("Điểm AI", 0) or 0),
                            "class_name": student_cache.get(student_id, {}).get("class_name", "C401"),
                            "trend": student_cache.get(student_id, {}).get("trend", "flat"),
                            "streak": student_cache.get(student_id, {}).get("streak", 0)
                        }
                        pitching_count += 1
            print(f"Enriched {pitching_count} pitching day students from CSV.")
        except Exception as e:
            print(f"Warning: Could not read pitching CSV: {e}")

    # 3. Read & parse Excel report file
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: Excel file not found at {EXCEL_PATH}")
        sys.exit(1)

    print(f"Loading workbook: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    
    raw_records = []
    
    # Read from Điểm cộng sheet if it exists, otherwise fall back to Form Responses 1
    if "Điểm cộng" in wb.sheetnames:
        print("Reading from sheet: 'Điểm cộng'")
        sheet = wb["Điểm cộng"]
        headers = None
        for row in sheet.iter_rows(values_only=True):
            if not headers:
                headers = row
                continue
            
            # Map columns
            session_id = row[headers.index("Session ID")]
            timestamp = row[headers.index("Thời gian gửi")]
            email = row[headers.index("Email")]
            raw_line = row[headers.index("Dòng gốc")]
            
            if not raw_line:
                continue
                
            parsed = parse_raw_line(str(raw_line))
            if parsed:
                raw_records.append({
                    "session_id": session_id,
                    "timestamp": timestamp,
                    "email": email,
                    **parsed
                })
    else:
        print("Sheet 'Điểm cộng' not found. Reading and splitting directly from 'Form Responses 1'...")
        sheet = wb["Form Responses 1"]
        headers = None
        for row in sheet.iter_rows(values_only=True):
            if not headers:
                headers = row
                continue
            
            session_id = row[headers.index("Mã buổi học / Session ID")]
            timestamp = row[headers.index("Timestamp")]
            email = row[headers.index("Email Address")]
            points_list_raw = row[headers.index("Danh sách học viên được cộng điểm")]
            
            if not points_list_raw:
                continue
                
            # Split multi-line reports by newlines
            lines = str(points_list_raw).split("\n")
            for line in lines:
                if not line.strip():
                    continue
                parsed = parse_raw_line(line)
                if parsed:
                    raw_records.append({
                        "session_id": session_id,
                        "timestamp": timestamp,
                        "email": email,
                        **parsed
                    })

    print(f"Parsed {len(raw_records)} point records from Excel.")

    # 4. Generate points_records for LabScoreLive
    points_records = []
    for idx, rec in enumerate(raw_records):
        # Format timestamp
        ts = rec["timestamp"]
        if isinstance(ts, datetime.datetime):
            ts_str = ts.strftime('%Y-%m-%d %H:%M:%S')
        else:
            ts_str = str(ts) if ts else datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Determine points type and category label
        p_type = "INDIVIDUAL"
        category_label = "👤 Điểm cá nhân phát biểu"
        reason_lower = rec["reason"].lower()
        
        if "hỗ trợ" in reason_lower:
            p_type = "SUPPORT"
            category_label = "🦸 Hỗ trợ đồng đội"
        elif "nhóm" in reason_lower or "chấm chéo" in reason_lower:
            p_type = "TEAM"
            category_label = "👥 Điểm Nhóm / Dự án"

        # Determine course and class info from session ID
        session_id = rec["session_id"] or "K3-DAY01-LEC-C401"
        parts = session_id.split("-")
        course_id = parts[0] if len(parts) > 0 else "K3"
        day = parts[1] if len(parts) > 1 else "DAY01"
        session_type = parts[2] if len(parts) > 2 else "LEC"
        class_name = parts[3] if len(parts) > 3 else "C401"

        # Check/update student cache with their class/course from points records
        student_id = rec["student_id"]
        student_name = rec["student_name"]
        
        if student_id not in student_cache:
            student_cache[student_id] = {
                "name": student_name,
                "course": "Khóa 4" if course_id == "K4" else "Khóa 3",
                "team_code": "N/A",
                "project_name": "Chưa có dự án",
                "role": "Thành viên",
                "ai_score": 0,
                "class_name": class_name,
                "trend": "flat",
                "streak": 0
            }
        else:
            # Update name if cached name has placeholder artifacts
            cached_s = student_cache[student_id]
            if len(student_name) > len(cached_s["name"]) or "+" in cached_s["name"]:
                cached_s["name"] = student_name
            # Ensure class name is set correctly
            if cached_s["class_name"] == "C401" and class_name != "C401":
                cached_s["class_name"] = class_name

        points_records.append({
            "id": f"PTS-{idx+1:04d}",
            "session_id": session_id,
            "class_name": class_name,
            "timestamp": ts_str,
            "coach_email": rec["email"] or "coach@labflow.edu.vn",
            "student_id": student_id,
            "student_name": student_name,
            "team_code": student_cache[student_id]["team_code"],
            "points": rec["points"],
            "type": p_type,
            "category_label": category_label,
            "reason": rec["reason"],
            "raw_line": rec["raw_line"],
            "course_id": course_id,
            "day": day,
            "session_type": session_type
        })

    # 5. Generate clean deduplicated students list
    students_list = []
    for s_id, s_info in student_cache.items():
        # Ensure name is properly clean and capitalized
        clean_name = clean_student_name(s_info["name"])
        students_list.append({
            "id": s_id,
            "name": clean_name,
            "course": s_info["course"],
            "team_code": s_info["team_code"],
            "project_name": s_info["project_name"],
            "role": s_info["role"],
            "ai_score": s_info["ai_score"],
            "class_name": s_info["class_name"],
            "trend": s_info["trend"],
            "streak": s_info["streak"]
        })

    # Sort students by name for presentation
    students_list.sort(key=lambda s: s["name"])

    # 6. Keep static sessions list
    sessions_list = [
        {"id": "K3-DAY01-LEC-D301", "name": "Buổi 1: Tổng quan & Nhập môn (D301)", "class_name": "D301", "date": "2026-07-23"},
        {"id": "K3-DAY01-LEC-C401", "name": "Buổi 1: Tổng quan & Nhập môn (C401)", "class_name": "C401", "date": "2026-07-23"},
        {"id": "K3-DAY02-LEC-D301", "name": "Buổi 2: Kỹ thuật Prompting (D301)", "class_name": "D301", "date": "2026-07-25"},
        {"id": "K3-DAY02-LEC-C401", "name": "Buổi 2: Kỹ thuật Prompting (C401)", "class_name": "C401", "date": "2026-07-25"},
        {"id": "K3-DAY03-LIVE-NOW", "name": "Buổi 3: [Đang Diễn Ra] Live Pitching & Demo (Toàn Trường)", "class_name": "ALL", "date": "2026-08-05", "is_live": True}
    ]

    output_data = {
        "students": students_list,
        "points_records": points_records,
        "sessions": sessions_list
    }

    # Write output JSON
    os.makedirs(os.path.dirname(JSON_PATH), exist_ok=True)
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)

    print(f"\nSuccessfully cleaned and output data to: {JSON_PATH}")
    print(f"Total clean students: {len(students_list)}")
    print(f"Total clean point records: {len(points_records)}")
    print("=== Cleanup Complete ===")

if __name__ == "__main__":
    main()
